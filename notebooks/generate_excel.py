import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side)
from openpyxl.utils import get_column_letter
import os

# ── LOAD ALL QUERY RESULTS ───────────────────────────────
region   = pd.read_csv("data/processed/query_results/revenue_and_profit_by_region.csv")
monthly  = pd.read_csv("data/processed/query_results/monthly_revenue_trend.csv")
products = pd.read_csv("data/processed/query_results/product_performance_by_sub_category.csv")
discount = pd.read_csv("data/processed/query_results/discount_impact_on_profit.csv")
segment  = pd.read_csv("data/processed/query_results/customer_segment_analysis.csv")
yoy      = pd.read_csv("data/processed/query_results/year_over_year_growth.csv")
top10    = pd.read_csv("data/processed/query_results/top_10_most_profitable_products.csv")
bottom10 = pd.read_csv("data/processed/query_results/bottom_10_loss_making_products.csv")

print("✅ All data loaded")

# ── CREATE EXCEL ─────────────────────────────────────────
os.makedirs("reports", exist_ok=True)
output_path = "reports/Sales_Performance_Analysis.xlsx"

with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
    region.to_excel(writer,   sheet_name="Regional Analysis",     index=False)
    monthly.to_excel(writer,  sheet_name="Monthly Trend",         index=False)
    products.to_excel(writer, sheet_name="Product Analysis",      index=False)
    discount.to_excel(writer, sheet_name="Discount Impact",       index=False)
    segment.to_excel(writer,  sheet_name="Customer Segments",     index=False)
    yoy.to_excel(writer,      sheet_name="YoY Growth",            index=False)
    top10["performance"]    = "TOP 10"
    bottom10["performance"] = "BOTTOM 10"
    combined = pd.concat([top10, bottom10], ignore_index=True)
    combined.to_excel(writer, sheet_name="Top & Bottom Products", index=False)

print("✅ Excel file created with 7 sheets")

# ── STYLE ────────────────────────────────────────────────
wb = load_workbook(output_path)

header_fill  = PatternFill("solid", fgColor="1F3864")
header_font  = Font(color="FFFFFF", bold=True, size=11)
border_side  = Side(style="thin", color="CCCCCC")
cell_border  = Border(left=border_side, right=border_side,
                      top=border_side,  bottom=border_side)
red_fill     = PatternFill("solid", fgColor="FFE0E0")
green_fill   = PatternFill("solid", fgColor="E0FFE0")

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]

    # Header row styling
    for cell in ws[1]:
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = cell_border

    # Data rows
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border    = cell_border
            cell.alignment = Alignment(horizontal="left", vertical="center")

    # Auto column width
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_length + 4, 40)

    # Freeze header
    ws.freeze_panes  = "A2"
    ws.row_dimensions[1].height = 25

# Highlight profits
for sheet_name in ["Regional Analysis", "Product Analysis",
                   "Discount Impact", "YoY Growth"]:
    ws = wb[sheet_name]
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                if cell.value < 0:
                    cell.fill = red_fill
                elif cell.value > 50:
                    cell.fill = green_fill

wb.save(output_path)
print(f"✅ Professional Excel saved → {output_path}")
print("✅ Open reports/Sales_Performance_Analysis.xlsx to view")